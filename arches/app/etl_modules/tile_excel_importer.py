from datetime import datetime
import json
from openpyxl import load_workbook
import os
from tempfile import NamedTemporaryFile

from django.core.exceptions import ValidationError
import uuid
from django.db import connection
from django.http import HttpRequest, HttpResponse
from django.utils.translation import gettext as _
from django.core.files.storage import default_storage
from django.contrib.auth.models import User
from arches.app.datatypes.datatypes import DataTypeFactory
from arches.app.etl_modules.decorators import load_data_async
from arches.app.models.models import (
    Node,
    TileModel,
    ETLModule,
    LoadEvent,
    LoadErrors,
    LoadStaging,
)
from arches.app.models.system_settings import settings
from arches.app.etl_modules.base_import_module import (
    BaseImportModule,
    FileValidationError,
)
import arches.app.tasks as tasks
from arches.management.commands.etl_template import create_tile_excel_workbook


class TileExcelImporter(BaseImportModule):
    def __init__(self, request=None, loadid=None, temp_dir=None, params=None):
        self.loadid = request.POST.get("load_id") if request else loadid
        self.userid = (
            request.user.id
            if request
            else settings.DEFAULT_RESOURCE_IMPORT_USER["userid"]
        )
        self.mode = "cli" if not request and params else "ui"
        try:
            self.user = User.objects.get(pk=self.userid)
        except User.DoesNotExist:
            raise User.DoesNotExist(
                _(
                    "The userid {} does not exist. Probably DEFAULT_RESOURCE_IMPORT_USER is not configured correctly in settings.py.".format(
                        self.userid
                    )
                )
            )
        if not request and params:
            request = HttpRequest()
            request.user = self.user
            request.method = "POST"
            for k, v in params.items():
                request.POST.__setitem__(k, v)
        self.request = request
        self.moduleid = request.POST.get("module") if request else None
        self.datatype_factory = DataTypeFactory()
        self.legacyid_lookup = {}
        self.validated_data = {}
        self.temp_path = ""
        self.temp_dir = temp_dir if temp_dir else None
        self.config = (
            ETLModule.objects.get(pk=self.moduleid).config if self.moduleid else {}
        )

    @load_data_async
    def run_load_task_async(self, request):
        self.loadid = request.POST.get("load_id")
        self.temp_dir = os.path.join(settings.UPLOADED_FILES_DIR, "tmp", self.loadid)
        self.file_details = request.POST.get("load_details", None)
        result = {}
        if self.file_details:
            details = json.loads(self.file_details)
            files = details["result"]["summary"]["files"]
            summary = details["result"]["summary"]

        load_task = tasks.load_tile_excel.apply_async(
            (self.userid, files, summary, result, self.temp_dir, self.loadid),
        )
        LoadEvent.objects.filter(loadid=self.loadid).update(taskid=load_task.task_id)

    def create_tile_value(
        self,
        data_node_lookup,
        node_lookup,
        nodegroup_alias,
        row_details,
    ):
        node_value_keys = data_node_lookup[nodegroup_alias]
        tile_value = {}
        tile_valid = True
        error_instances = []
        for key in node_value_keys:
            try:
                nodeid = node_lookup[key]["nodeid"]
                node_details = node_lookup[key]
                datatype = node_details["datatype"]
                datatype_instance = self.datatype_factory.get_instance(datatype)
                source_value = row_details[key]
                config = node_details["config"]

                config["bulk_import"] = True

                config["loadid"] = self.loadid
                try:
                    config["nodeid"] = nodeid
                except TypeError:
                    config = {}

                value, validation_errors = self.prepare_data_for_loading(
                    datatype_instance, source_value, config
                )
                valid = True if len(validation_errors) == 0 else False
                if not valid:
                    tile_valid = False
                error_message = ""
                for error in validation_errors:
                    error_message = (
                        "{0}|{1}".format(error_message, error["message"])
                        if error_message != ""
                        else error["message"]
                    )
                    error_instances.append(
                        LoadErrors(
                            type="node",
                            value=str(source_value),
                            source="",
                            error=error["title"],
                            message=error["message"],
                            datatype=datatype,
                            load_event_id=self.loadid,
                            node_id=nodeid,
                        )
                    )

                if value is not None:
                    tile_value[nodeid] = {
                        "value": value,
                        "valid": valid,
                        "source": source_value,
                        "notes": error_message,
                        "datatype": datatype,
                    }
                else:
                    tile_value[nodeid] = None
            except KeyError:
                pass

        return tile_value, tile_valid, error_instances

    def get_nodegroup_id_column(self, worksheet):
        """
        Returns the index of the column that contains the nodegroup id.
        If no nodegroup id is found, returns None.
        """
        index = 1
        for row in worksheet.iter_rows(1, 1, None, None):
            for cell in row:
                if cell.value == "nodegroup_id":
                    return index
                else:
                    index += 1
        return worksheet.max_column

    def process_worksheet(self, worksheet, cursor, node_lookup, nodegroup_lookup):
        data_node_lookup = {}
        row_count = 0
        nodegroupid_column = self.get_nodegroup_id_column(worksheet)
        maybe_nodegroup = worksheet.cell(row=2, column=nodegroupid_column).value
        if maybe_nodegroup:
            nodegroup_alias = nodegroup_lookup[maybe_nodegroup]["alias"]
            data_node_lookup[nodegroup_alias] = [
                val.value for val in worksheet[1][3:-3]
            ]

        staging_instances = []
        all_error_instances = []

        tiles_to_update = []

        for row in worksheet.iter_rows(min_row=2):
            cell_values = [cell.value for cell in row]
            if len(cell_values) == 0 or any(cell_values) is False:
                continue
            resourceid = cell_values[2]
            if resourceid is None:
                LoadEvent.objects.filter(loadid=self.loadid).update(
                    status="failed", load_end_time=datetime.now()
                )
                raise ValueError(_("All rows must have a valid resource id"))

            node_values = cell_values[3:-3]
            sortorder = cell_values[-3] if cell_values[-3] else 0
            try:
                row_count += 1
                row_details = dict(zip(data_node_lookup[nodegroup_alias], node_values))
                row_details["nodegroup_id"] = node_lookup[nodegroup_alias]["nodeid"]
                user_tileid = (
                    cell_values[0].strip()
                    if cell_values[0] and cell_values[0] != "None"
                    else None
                )
                tileid = user_tileid if user_tileid else uuid.uuid4()
                nodegroup_depth = nodegroup_lookup[row_details["nodegroup_id"]]["depth"]
                parenttileid = (
                    cell_values[1].strip()
                    if cell_values[1] and cell_values[1] != "None"
                    else None
                )
                legacyid, resourceid = self.set_legacy_id(resourceid)
                tile_value, passes_validation, error_instances = self.create_tile_value(
                    data_node_lookup,
                    node_lookup,
                    nodegroup_alias,
                    row_details,
                )
                all_error_instances.extend(error_instances)
                nodegroup_cardinality = nodegroup_lookup[row_details["nodegroup_id"]][
                    "cardinality"
                ]
                operation = "insert"
                if user_tileid:
                    if nodegroup_cardinality == "n":
                        operation = (
                            "update"  # db will "insert" if tileid does not exist
                        )
                    elif nodegroup_cardinality == "1":
                        operation = "insert"
                        tiles_to_update.append((len(staging_instances), tileid))
                staging_instances.append(
                    LoadStaging(
                        nodegroup_id=row_details["nodegroup_id"],
                        legacyid=legacyid,
                        resourceid=resourceid,
                        tileid=tileid,
                        parenttileid=parenttileid,
                        value=tile_value,
                        load_event_id=self.loadid,
                        nodegroup_depth=nodegroup_depth,
                        source_description="worksheet:{0}, row:{1}".format(
                            worksheet.title, row[0].row
                        ),
                        passes_validation=passes_validation,
                        operation=operation,
                        sortorder=sortorder,
                    )
                )
            except KeyError:
                pass

        if tiles_to_update:
            pending_tileids = [t for _, t in tiles_to_update]
            existing_tileids = set(
                TileModel.objects.filter(pk__in=pending_tileids).values_list(
                    "pk", flat=True
                )
            )
            for staging_instances_idx, tileid in tiles_to_update:
                if tileid in existing_tileids:
                    staging_instances[staging_instances_idx].operation = "update"

        batch_size = settings.BULK_IMPORT_BATCH_SIZE
        LoadErrors.objects.bulk_create(all_error_instances, batch_size=batch_size)
        LoadStaging.objects.bulk_create(staging_instances, batch_size=batch_size)
        cursor.execute(
            """
            INSERT INTO load_errors (type, source, error, loadid, nodegroupid)
            SELECT 'tile', source_description, error_message, loadid, nodegroupid
            FROM load_staging
            WHERE loadid = %s AND passes_validation = false AND error_message IS NOT null
            """,
            [self.loadid],
        )
        return {"name": worksheet.title, "rows": row_count}

    def validate_uploaded_file(self, workbook):
        graphid = None
        for worksheet in workbook.worksheets:
            if worksheet.cell(2, self.get_nodegroup_id_column(worksheet)).value:
                try:
                    nodegroup_id = worksheet.cell(
                        2, self.get_nodegroup_id_column(worksheet)
                    ).value
                    graphid = str(
                        Node.objects.filter(nodegroup_id=nodegroup_id)[0].graph_id
                    )
                    break
                except (IndexError, ValidationError):
                    pass
        if graphid is None:
            raise FileValidationError()

    def get_graphid(self, workbook):
        for worksheet in workbook.worksheets:
            if worksheet.cell(2, self.get_nodegroup_id_column(worksheet)).value:
                try:
                    nodegroup_id = worksheet.cell(
                        2, self.get_nodegroup_id_column(worksheet)
                    ).value
                    graphid = str(
                        Node.objects.filter(nodegroup_id=nodegroup_id)[0].graph_id
                    )
                    break
                except (IndexError, ValidationError):
                    pass
        return graphid

    def stage_files(self, files, summary, cursor):
        for file in files:
            self.stage_excel_file(file, summary, cursor)

    def stage_excel_file(self, file, summary, cursor):
        if file.endswith("xlsx") and ("attachments" + os.sep) not in file:
            summary["files"][file]["worksheets"] = []
            uploaded_file_path = os.path.join(
                settings.UPLOADED_FILES_DIR, "tmp", self.loadid, file
            )
            opened_file = default_storage.open(uploaded_file_path)
            workbook = load_workbook(filename=opened_file, read_only=True)
            graphid = self.get_graphid(workbook)
            nodegroup_lookup, nodes = self.get_graph_tree(graphid)
            node_lookup = self.get_node_lookup(nodes)

            for worksheet in workbook.worksheets:
                details = self.process_worksheet(
                    worksheet, cursor, node_lookup, nodegroup_lookup
                )
                summary["files"][file]["worksheets"].append(details)
            opened_file.close()
            LoadEvent.objects.filter(loadid=self.loadid).update(load_details=summary)

    def download(self, request):
        format = request.POST.get("format")
        if format == "xls":
            wb = create_tile_excel_workbook(request.POST.get("id"))
            with NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                wb.save(tmp.name)
                tmp.seek(0)
                response = HttpResponse(
                    tmp.read(), content_type="application/vnd.ms-excel"
                )
                response["Content-Disposition"] = "attachment"
            os.unlink(tmp.name)
            return {"success": True, "raw": response}
        else:
            return {"success": False, "data": "failed"}
